import refinitiv.data as rd
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import uuid
# from builtins import print,int,str,enumerate,len,all,set,float,any,list,ValueError

# Initialize Refinitiv session
rd.open_session()

# Configuration
poa_input = "CY2026"
poa_type = poa_input[:2]
poa_year = int(poa_input[2:])
companies = ['CRDA.L']
scale = 6
pod_cutoff_estimate = pd.to_datetime("2024-11-01").date()
today = pd.to_datetime("today").normalize().date()
cutoff_date_POA = today + timedelta(days=3650)

# Column names (dynamic)
col_ticker = "Ticker"
col_broker_name = "Broker Name"
col_analyst_name = "Analyst Name"
col_estimate_date = "Estimate Date"
col_target_date = "Target Date"
col_target_price = "Broker Target"
col_dps = "DPS"
col_div_yield = "Dividend Yield"
col_div_yield_date = "Dividend Yield Date"
col_ev = "EV"
col_ebitda = "EBITDA"
col_ebitda_margin = "EBITDA Margin"
col_ebit_margin = "EBIT Margin"
col_net_debt = "Net Debt"
col_shares = "Shares Outstanding"
col_rec_label = "Recommendation"
col_rec_date = "Recommendation Date"
col_revenue = "Revenue"
col_rev_date = "Revenue Date"
col_ebitda_date = "EBITDA Date"
col_net_debt_date = "Net Debt Date"
col_shares_date = "No. of Shares Outstanding Date"
col_price = "Price"
col_market_cap = "Market Cap"
col_ev_ebitda = "EV/EBITDA"
col_ebit = "EBIT"
col_ebitda_12m_fwd = "EBITDA (12M Fwd)"

# Broker overrides
refinitiv_override = {
    "PERMISSION DENIED 1342152": "SBI SECURITIES",
    "PERMISSION DENIED 87408": "ROBERT W. BAIRD & CO",
    "PERMISSION DENIED 937880": "TACHIBANA SECURITIES",
    "PERMISSION DENIED 1120": "JEFFERIES",
    "PERMISSION DENIED 1207112": "MELIUS RESEARCH",
    "PERMISSION DENIED 1424952": "CFRA RESEARCH",
    "PERMISSION DENIED 156648": "IWAICOSMO SECURITIES",
    "PERMISSION DENIED 17472": "RBC CAPITAL MARKETS",
    "PERMISSION DENIED 211744": "CROSS RESEARCH",
    "PERMISSION DENIED 22760": "CLSA",
    "PERMISSION DENIED 23440": "MIZUHO",
    "PERMISSION DENIED 23816": "MORGAN STANLEY",
    "PERMISSION DENIED 25632": "CITIGROUP",
    "PERMISSION DENIED 266912": "KEPLER CHEUVREUX",
    "PERMISSION DENIED 284328": "ARETE RESEARCH SERVICES LLP",
    "PERMISSION DENIED 2880": "HSBC",
    "PERMISSION DENIED 310016": "REDBURN ATLANTIC",
    "PERMISSION DENIED 32": "BOFA",
    "PERMISSION DENIED 32848": "BMO CAPITAL",
    "PERMISSION DENIED 347360": "WOLFE RESEARCH",
    "PERMISSION DENIED 36928": "JP MORGAN",
    "PERMISSION DENIED 392": "DEUTSCHE BANK",
    "PERMISSION DENIED 398136": "BARCLAYS",
    "PERMISSION DENIED 483808": "HAITONG INTERNATIONAL",
    "PERMISSION DENIED 495296": "MIZUHO",
    "PERMISSION DENIED 512368": "SMBC NIKKO",
    "PERMISSION DENIED 54992": "BERNSTEIN",
    "PERMISSION DENIED 662336": "ALPHAVALUE",
    "PERMISSION DENIED 696": "TD COWEN",
    "PERMISSION DENIED 73704": "GOLDMAN SACHS",
    "PERMISSION DENIED 7896": "DAIWA SECURITIES",
    "PERMISSION DENIED 85152": "CANACCORD GENUITY"
}

def format_dates(df):
    date_columns = [col for col in df.columns if "Date" in col]
    for col in date_columns:
        df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d %b %y")
    return df

def consolidate_refinitiv_data(df, key_columns=None):
    if key_columns is None:
        key_columns = [col_ticker, col_broker_name]
        if col_estimate_date in df.columns:
            key_columns.append(col_estimate_date)
    
    missing_cols = [col for col in key_columns if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Key column(s) {missing_cols} not found in dataframe")
    
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    
    for col in df.columns:
        if col not in numeric_cols:
            try:
                temp = pd.to_numeric(df[col], errors='coerce')
                if temp.notna().mean() > 0.5:
                    df[col] = temp
                    numeric_cols.append(col)
            except:
                pass
    
    metadata_cols = [col for col in df.columns if col not in key_columns and col not in numeric_cols]
    
    aggregations = {}
    for col in numeric_cols:
        aggregations[col] = lambda x: x.dropna().iloc[0] if not x.dropna().empty else np.nan
    for col in metadata_cols:
        aggregations[col] = 'first'
    
    result = df.groupby(key_columns, as_index=False).agg(aggregations)
    return result

def apply_broker_overrides(df):
    if col_broker_name in df.columns:
        df[col_broker_name] = df[col_broker_name].astype(str).str.upper().str.strip()
        for denied_key, broker_name in refinitiv_override.items():
            df[col_broker_name] = df[col_broker_name].replace(denied_key, broker_name)
    return df

def get_metric_cy(metric_code, label, scale_on=True):
    scale_str = f",Scale={scale}" if scale_on else ""
    df = rd.get_data(
        universe=companies,
        fields=[f"{metric_code}.brokername;{metric_code}.date;{metric_code}{scale_str}"],
        parameters={"Period": poa_input}
    )
    df.columns = [col_ticker, col_broker_name, col_estimate_date, label]
    
    df = apply_broker_overrides(df)
    df[col_estimate_date] = pd.to_datetime(df[col_estimate_date], errors="coerce").dt.date
    df = df[(df[col_estimate_date] >= pod_cutoff_estimate)]
    return df.dropna(subset=[col_broker_name, col_estimate_date, label])

def get_metric_fy(metric_code, label, scale_on=True):
    scale_str = f",Scale={scale}" if scale_on else ""
    df = rd.get_data(
        universe=companies,
        fields=[f"{metric_code}.brokername;{metric_code}.date;{metric_code}{scale_str}"],
        parameters={"Period": poa_input}
    )
    df.columns = [col_ticker, col_broker_name, col_estimate_date, label]
    
    df = apply_broker_overrides(df)
    df[col_estimate_date] = pd.to_datetime(df[col_estimate_date], errors="coerce").dt.date
    df = df[(df[col_estimate_date] >= pod_cutoff_estimate)]
    return df.dropna(subset=[col_broker_name, col_estimate_date, label])

def get_estimate_date(metric_date_field, label):
    df = rd.get_data(
        universe=companies,
        fields=[f"{metric_date_field}.brokername;{metric_date_field}.date"],
        parameters={"Period": poa_input}
    )
    df.columns = [col_ticker, col_broker_name, label]
    
    df = apply_broker_overrides(df)
    df[label] = pd.to_datetime(df[label], errors="coerce").dt.date
    return df.dropna(subset=[col_broker_name, label])

def create_multi_metric_forecast_summary(df, metrics, output_file="Multi_Metric_Forecast_Summary.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Summary"
    
    # Define the statistical measures for the summary table
    stats_measures = [
        "Median",
        "10th Percentile",
        "90th Percentile"
    ]
    
    current_row = 1
    tickers = df[col_ticker].unique()
    
    for ticker in tickers:
        ticker_data = df[df[col_ticker] == ticker].copy()
        
        # Write the forecast panel header
        ws.cell(row=current_row, column=1, value=f"{ticker} FORECAST PANEL")
        current_row += 1
        
        # Define columns for the forecast panel
        display_cols = [
            col_ticker, col_broker_name, col_analyst_name,
            f"{col_revenue} {poa_input}", f"{col_ebitda} {poa_input}",
            col_price, f"{col_net_debt} {poa_input}", f"{col_shares} {poa_input}",
            f"{col_ebitda_margin} {poa_input}", f"{col_ev_ebitda} {poa_input}",
            f"{poa_input} {col_div_yield}", col_ebitda_12m_fwd
        ]
        valid_cols = [col for col in display_cols if col in df.columns]
        
        # Write forecast panel headers
        for col_idx, header in enumerate(valid_cols, 1):
            ws.cell(row=current_row, column=col_idx, value=header)
        current_row += 1
        
        # Write forecast panel data and track the row range
        start_data_row = current_row
        for _, row in ticker_data.iterrows():
            for col_idx, header in enumerate(valid_cols, 1):
                value = row.get(header, None)
                if pd.isna(value):
                    value = None
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
        end_data_row = current_row - 1
        
        # Add one-row gap after forecast panel
        current_row += 1
        
        # Write summary table header
        ws.cell(row=current_row, column=1, value=f"Summary Statistics - {ticker}")
        current_row += 1
        
        # Write summary table column headers
        ws.cell(row=current_row, column=1, value="Statistic")
        for col_idx, metric in enumerate(metrics, 2):
            ws.cell(row=current_row, column=col_idx, value=metric)
        current_row += 1
        
        # Write summary table
        for stat_idx, stat in enumerate(stats_measures, 1):
            ws.cell(row=current_row, column=1, value=stat)
            for col_idx, metric in enumerate(metrics, 2):
                if metric not in valid_cols:
                    continue
                # Extract the data for numpy calculations
                values = ticker_data.get(metric, pd.Series()).dropna()
                try:
                    values = values.astype(float)
                except:
                    ws.cell(row=current_row, column=col_idx, value=None)
                    continue
                
                # Filter out zeros to match IF(<range><>0, ...)
                filtered_values = values[values != 0]
                if len(filtered_values) == 0:
                    value = None
                else:
                    if stat == "Median":
                        value = np.median(filtered_values)
                    elif stat == "10th Percentile":
                        value = np.percentile(filtered_values, 10)
                    elif stat == "90th Percentile":
                        value = np.percentile(filtered_values, 90)
                    else:
                        value = None
                # Write the static value
                ws.cell(row=current_row, column=col_idx, value=value)
                
                # Apply percentage format for specific metrics
                if any(m in metric for m in ["Margin", "Dividend Yield"]):
                    ws.cell(row=current_row, column=col_idx).number_format = '0.0%'
            current_row += 1
        
        # Add two-row gap after summary table (unless it's the last ticker)
        if ticker != tickers[-1]:
            current_row += 2
    
    wb.save(output_file)
    print(f"Multi-metric forecast summary with numpy percentiles saved to {output_file}")
    
    # Return a dictionary of DataFrames for verification
    summary_dfs = {}
    for ticker in tickers:
        ticker_data = df[df[col_ticker] == ticker].copy()
        forecast_df = ticker_data[valid_cols]
        
        summary_data = []
        for stat in stats_measures:
            stat_row = {"Statistic": stat}
            for metric in metrics:
                values = ticker_data.get(metric, pd.Series()).dropna()
                try:
                    values = values.astype(float)
                except:
                    stat_row[metric] = None
                    continue
                if len(values) == 0:
                    stat_row[metric] = None
                    continue
                    
                # Filter out zeros for percentiles
                filtered_values = values[values != 0]
                if len(filtered_values) == 0:
                    stat_row[metric] = None
                    continue
                if stat == "Median":
                    value = np.median(filtered_values)
                elif stat == "10th Percentile":
                    value = np.percentile(filtered_values, 10)
                elif stat == "90th Percentile":
                    value = np.percentile(filtered_values, 90)
                else:
                    value = None
                
                if value is not None:
                    if any(m in metric for m in ["Margin", "Dividend Yield"]):
                        value = value * 100
                    stat_row[metric] = value
                else:
                    stat_row[metric] = None
            summary_data.append(stat_row)
        
        summary_dfs[ticker] = {
            "Forecast Panel": forecast_df,
            "Summary": pd.DataFrame(summary_data)
        }
    
    return summary_dfs

# Main data processing
metrics = {
    f"{col_revenue} {poa_input}": "TR.RevenueEstValue",
    f"{col_ebitda} {poa_input}": "TR.EBITDAEstValue",
    f"{col_ebit} {poa_input}": "TR.EBITEstValue",
    f"{col_net_debt} {poa_input}": "TR.NetDebtEstValue",
    f"{col_dps} {poa_input}": "TR.DPSEstValue",
    f"{col_ev} {poa_input}": "TR.EVEstValue"
}

raw_data_frames = {}
data_frames = []
for label, code in metrics.items():
    if poa_type == "CY":
        df = get_metric_cy(code, label, scale_on=False if col_dps in label else True)
    else:
        df = get_metric_fy(code, label, scale_on=False if col_dps in label else True)
    data_frames.append(df)
    raw_data_frames[label] = df.copy()

# Shares data
shares_df = rd.get_data(
    universe=companies,
    fields=[f"TR.NumberOfSharesOutstanding.brokername;TR.NumberOfSharesOutstanding(Period={poa_input})"]
)
shares_df.columns = [col_ticker, col_broker_name, f"{col_shares} {poa_input}"]
shares_df = apply_broker_overrides(shares_df)
shares_df = shares_df.dropna(subset=[col_broker_name, f"{col_shares} {poa_input}"]).drop_duplicates(subset=[col_ticker, col_broker_name])
data_frames.append(shares_df)
raw_data_frames[f"{col_shares} {poa_input}"] = shares_df.copy()

# Price data
price_df = rd.get_data(
    universe=companies,
    fields=["TR.PriceClose"]
)
price_df.columns = [col_ticker, col_price]
data_frames.append(price_df)

# Recommendation data
rec_df = rd.get_data(
    universe=companies,
    fields=["TR.BrkRecEstBrokerName", "TR.BrkRecLabel", "TR.BrkRecLabelEstDate"],
    parameters={"Period": poa_input}
)
rec_df.columns = [col_ticker, col_broker_name, col_rec_label, col_rec_date]
rec_df = apply_broker_overrides(rec_df)
rec_df[col_rec_date] = pd.to_datetime(rec_df[col_rec_date], errors="coerce").dt.date
rec_df = rec_df.drop_duplicates(subset=[col_ticker, col_broker_name])
data_frames.append(rec_df)
raw_data_frames[col_rec_label] = rec_df.copy()

# Target price data
tp_df = rd.get_data(
    universe=companies,
    fields=["TR.TPEstValue.brokername;TR.TPEstValue.date;TR.TPEstValue;TR.AnalystName"],
    parameters={"Period": poa_input}
)
tp_df.columns = [col_ticker, col_broker_name, col_target_date, col_target_price, col_analyst_name]
tp_df = apply_broker_overrides(tp_df)
tp_df[col_target_date] = pd.to_datetime(tp_df[col_target_date], errors="coerce").dt.date
tp_df = tp_df.drop_duplicates(subset=[col_ticker, col_broker_name])
data_frames.append(tp_df)
raw_data_frames[col_target_price] = tp_df.copy()

# Date fields
revenue_date_df = get_estimate_date("TR.RevenueEstDate", f"{poa_input} {col_rev_date}")
ebitda_date_df = get_estimate_date("TR.EBITDAEstDate", f"{poa_input} {col_ebitda_date}")
netdebt_date_df = get_estimate_date("TR.NetDebtEstDate", f"{poa_input} {col_net_debt_date}")
shares_date_df = get_estimate_date("TR.NumberOfSharesOutstanding", f"{poa_input} {col_shares_date}")
data_frames.extend([revenue_date_df, ebitda_date_df, netdebt_date_df, shares_date_df])

# Consolidate data
all_tickers_brokers = pd.DataFrame()
for df in data_frames:
    if col_broker_name in df.columns and col_ticker in df.columns:
        temp_df = df[[col_ticker, col_broker_name]].drop_duplicates()
        all_tickers_brokers = pd.concat([all_tickers_brokers, temp_df])
all_tickers_brokers = all_tickers_brokers.drop_duplicates()

panel = all_tickers_brokers.copy()
for df in data_frames:
    if col_broker_name not in df.columns:
        if col_ticker in df.columns:
            panel = pd.merge(panel, df, on=col_ticker, how="left")
    else:
        common_cols = list(set([col_ticker, col_broker_name]).intersection(df.columns))
        panel = pd.merge(panel, df, on=common_cols, how="left", suffixes=('', '_drop'))

panel = panel[[col for col in panel.columns if not col.endswith('_drop')]]

key_columns = [col_ticker, col_broker_name]
if col_estimate_date in panel.columns:
    key_columns.append(col_estimate_date)

panel = consolidate_refinitiv_data(panel, key_columns=key_columns)

# Calculate derived metrics
if all(col in panel.columns for col in [f"{col_ebitda} {poa_input}", f"{col_revenue} {poa_input}"]):
    panel[f"{col_ebitda_margin} {poa_input}"] = panel[f"{col_ebitda} {poa_input}"] / panel[f"{col_revenue} {poa_input}"]
if all(col in panel.columns for col in [f"{col_ebit} {poa_input}", f"{col_revenue} {poa_input}"]):
    panel[f"{col_ebit_margin} {poa_input}"] = panel[f"{col_ebit} {poa_input}"] / panel[f"{col_revenue} {poa_input}"]
if all(col in panel.columns for col in [f"{col_shares} {poa_input}", col_price]):
    panel[f"{col_market_cap} {poa_input}"] = panel[f"{col_shares} {poa_input}"] * panel[col_price]
if all(col in panel.columns for col in [f"{col_ev} {poa_input}", f"{col_ebitda} {poa_input}"]):
    panel[f"{col_ev_ebitda} {poa_input}"] = panel[f"{col_ev} {poa_input}"] / panel[f"{col_ebitda} {poa_input}"]
if all(col in panel.columns for col in [f"{col_dps} {poa_input}", col_price]):
    panel[f"{poa_input} {col_div_yield}"] = panel[f"{col_dps} {poa_input}"] / panel[col_price]
    panel[f"{poa_input} {col_div_yield_date}"] = today

# Assume EBITDA (12M Fwd) is the same as EBITDA for this example; adjust if different data source
panel[col_ebitda_12m_fwd] = panel[f"{col_ebitda} {poa_input}"]

# Format dates
panel = format_dates(panel)

# Define metrics for the summary table
metrics_to_analyze = [
    f"{col_revenue} {poa_input}",
    f"{col_ebitda} {poa_input}",
    col_price,
    f"{col_net_debt} {poa_input}",
    f"{col_shares} {poa_input}",
    f"{col_ebitda_margin} {poa_input}",
    f"{col_ev_ebitda} {poa_input}",
    f"{poa_input} {col_div_yield}",
    col_ebitda_12m_fwd
]

# Generate the combined forecast and summary sheet
summary_dfs = create_multi_metric_forecast_summary(panel, metrics_to_analyze, output_file="Combined_Forecast_Summary_With_Linking.xlsx")

# Print the DataFrames for verification
for ticker, dfs in summary_dfs.items():
    print(f"\nForecast Panel for {ticker}:")
    print(dfs["Forecast Panel"])
    # Define number of simulations
    n_simulations = 10000

    # Extract median, 10th and 90th percentiles for Revenue, EBITDA Margin and EV/EBITDA
    stats = {}
    for metric in [f"{col_revenue} {poa_input}", f"{col_ebitda_margin} {poa_input}", f"{col_ev_ebitda} {poa_input}"]:
        values = dfs["Forecast Panel"][metric].dropna()
        if not values.empty:
            stats[metric] = {
                'median': values.median(),
                'p10': values.quantile(0.1),
                'p90': values.quantile(0.9)
            }
            # Print percentiles for the current metric
            print(f"\n{metric} for {ticker}:")
            print(f"10th percentile: {stats[metric]['p10']:.2f}{'%' if 'Margin' in metric or 'Dividend Yield' in metric else ''}")
            print(f"Median: {stats[metric]['median']:.2f}{'%' if 'Margin' in metric or 'Dividend Yield' in metric else ''}")
            print(f"90th percentile: {stats[metric]['p90']:.2f}{'%' if 'Margin' in metric or 'Dividend Yield' in metric else ''}")
    print(f"\nSummary Table for {ticker}:")
    print(dfs["Summary"])